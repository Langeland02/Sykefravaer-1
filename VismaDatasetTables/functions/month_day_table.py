import os

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule


def create_matplotlib_bar_chart(df, title, output_filename):
    """
    Creates a bar chart using Matplotlib based on the monthly sums in the dataframe.
    """
    # Extract the monthly sums from the last row of the pivot table (excluding the 'Sum' column)
    months = df.columns[:-1]  # Excluding the 'Sum' column
    sums = df.loc['Sum', months]

    # Create the bar chart with specified figure size
    plt.figure(figsize=(7, 5))  # 3 inches by 5 inches corresponds to 300x500 pixels with 100 dpi
    bars = plt.bar(months, sums, color='#FF9999', edgecolor='#5A5A5A')  # Salmon red bars with dark-grey borders

    # Add data labels on top of each bar
    for bar in bars:
        yval = bar.get_height()
        plt.text(bar.get_x() + bar.get_width() / 2, yval + 0.5, int(yval), ha='center', va='bottom', fontsize=10)

    # Customize the chart: remove axis titles and give more space above the highest bar
    plt.title(title, fontsize=12)  # Title with fontsize 12
    plt.grid(axis='y', linestyle='--', alpha=0.7)

    # Keep x-axis labels (month names) visible
    plt.xticks(fontsize=10, rotation=45, ha='right')  # Rotate month names for better readability

    # Hide y-axis labels but keep the ticks
    plt.gca().axes.get_yaxis().set_visible(False)

    # Adjust the y-limits to add space above the highest bar
    plt.ylim(0, max(sums) * 1.2)

    plt.tight_layout()

    # Save the chart as an image and insert it into the Excel file
    chart_image_path = output_filename.replace('.xlsx', '_chart.png')
    plt.savefig(chart_image_path, dpi=100)  # Set dpi to 100 to match the desired pixel size (3"x5" = 300x500)
    plt.close()  # Close the plot to avoid display in non-interactive environments

    return chart_image_path
def summarize_day_month_counts(df, dayName_col, monthName_col, count_col, search_string, title, note, output_filename='tables/year_summary.xlsx'):
    # Define day and month order in Norwegian with only the first letter uppercase
    day_order = ['Mandag', 'Tirsdag', 'Onsdag', 'Torsdag', 'Fredag', 'Lørdag', 'Søndag']
    month_order = ['Januar', 'Februar', 'Mars', 'April', 'Mai', 'Juni', 'Juli', 'August', 'September', 'Oktober', 'November', 'Desember']

    # Avoid the SettingWithCopyWarning by using .loc[]
    df = df.copy()  # Make a copy to prevent modifying the original DataFrame

    # Convert the day and month columns to categorical types with the specified order
    df.loc[:, dayName_col] = pd.Categorical(df[dayName_col].str.capitalize(), categories=day_order, ordered=True)
    df.loc[:, monthName_col] = pd.Categorical(df[monthName_col].str.capitalize(), categories=month_order, ordered=True)

    # Filter rows where search_string appears in the specified column
    filtered_df = df[df[count_col].str.contains(search_string, na=False, case=False)]

    # Group by day name and month name, then count the occurrences
    summary_df = filtered_df.groupby([dayName_col, monthName_col], observed=True).size().reset_index(name='count')

    # Pivot the data to have months as columns and days as rows
    pivot_df = summary_df.pivot(index=dayName_col, columns=monthName_col, values='count').fillna(0)

    # Reorder both the rows (days) and columns (months) to ensure they are displayed in the correct order
    pivot_df = pivot_df.reindex(index=day_order, columns=month_order, fill_value=0)

    # Add a sum row at the bottom of the pivot table for each column (month)
    pivot_df.loc['Sum'] = pivot_df.sum()

    # Add a sum column at the end of the pivot table for each row (day)
    pivot_df['Sum'] = pivot_df.sum(axis=1)

    # Save the pivoted and reordered DataFrame to an Excel file
    pivot_df.to_excel(output_filename, index=True)

    # Load the workbook to apply formatting
    wb = load_workbook(output_filename)
    ws = wb.active

    # Insert two empty rows at the top to make space for the title
    ws.insert_rows(1, amount=2)

    # Add the title in cell A1 and merge it across the width of the table, aligned to the left
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    ws['A1'].value = title
    ws['A1'].font = Font(name='Times New Roman', bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')

    # Set Times New Roman for all cells and left-align text in column A
    for row in ws.iter_rows(min_row=3, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = Font(name='Times New Roman')
            if cell.column == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # Remove content of A3 (formerly A1, now shifted)
    ws['A3'].value = ''

    # Remove all borders
    no_border = Border(left=Side(border_style=None), right=Side(border_style=None),
                       top=Side(border_style=None), bottom=Side(border_style=None))
    for row in ws.iter_rows(min_row=3, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = no_border

    # Add a border under the header row (now row 3)
    thin_border = Border(bottom=Side(style='thin'))
    for cell in ws[3]:
        cell.border = thin_border

    # Add a border between columns A and B
    for row in ws.iter_rows(min_row=3, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
        ws[get_column_letter(2) + str(row[0].row)].border = Border(left=Side(style='thin'))

    # Add a border below "Januar" in column B (this is now in cell B3)
    ws['B3'].border = Border(bottom=Side(style='thin'))

    # Apply conditional formatting for a color scale from white to red based on the cell values
    color_scale_rule = ColorScaleRule(
        start_type='min', start_color='FFFFFF',
        mid_type='percentile', mid_value=50, mid_color='FF9999',
        end_type='max', end_color='FF0000'
    )

    # Apply the rule to the range of cells with numbers
    ws.conditional_formatting.add(f"B4:{get_column_letter(ws.max_column - 1)}{ws.max_row - 1}", color_scale_rule)

    # Add the note in rows 13, 14, and 15
    ws.merge_cells(start_row=13, start_column=1, end_row=15, end_column=ws.max_column)
    ws['A13'].value = note
    ws['A13'].font = Font(name='Times New Roman', size=8)
    ws['A13'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Ensure the background color for rows 1, 2, 12-15 is white
    for row in [1, 2, 12, 13, 14, 15]:
        for cell in ws[row]:
            cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # Add a border between rows 10 and 11
    for cell in ws[11]:
        cell.border = Border(top=Side(style='thin'))

    # Add a border between columns M and N (Desember and Sum) for rows 3 through 11
    for row in range(3, 12):
        existing_border = ws[f'M{row}'].border
        ws[f'M{row}'].border = Border(
            left=existing_border.left,
            right=Side(style='thin'),
            top=existing_border.top,
            bottom=existing_border.bottom
        )

    # Make the cell N11 bold
    ws['N11'].font = Font(name='Times New Roman', bold=True)

    # Create the Matplotlib bar chart and get the image path
    chart_image_path = create_matplotlib_bar_chart(pivot_df, title, output_filename)

    # Insert the chart image into the Excel file (at cell O1)
    img = Image(chart_image_path)
    ws.add_image(img, 'A16')

    # Save the formatted workbook
    wb.save(output_filename)
    print(f"Summary table saved to {output_filename}")

# Example usage
# df = pd.DataFrame({...})
# summarize_day_month_counts(df, 'day', 'month', 'status', 'completed', 'Your Title Here', 'Your note here.')
